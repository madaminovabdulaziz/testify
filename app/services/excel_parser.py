"""Excel test-upload parser.

PRODUCT_BLUEPRINT §12 defines the input contract; this module is the
gatekeeper that converts a teacher-edited ``.xlsx`` into either a
validated :class:`ParsedTest` or a list of :class:`ParseError` that the
admin can use to fix the file. **Errors are data, not exceptions** — the
caller branches on the return type (ARCHITECTURE_SPEC §8.7).

This module performs *no* DB writes. The caller persists ``ParsedTest``
through ``TestService`` / ``QuestionRepository.bulk_create``.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Final

from openpyxl import load_workbook

from app.services.question_validation import (
    EXPECTED_ROW_COUNT,
    MAX_IMAGE_CAPTION_BLOCK_LEN,
    MAX_OPTION_TEXT_LEN,
    MAX_QUESTION_TEXT_LEN,
    SECTION_RANGES,
    VALID_OPTIONS,
    VALID_SECTIONS,
    validate_question_fields,
)

__all__ = [
    "EXPECTED_ROW_COUNT",
    "MAX_IMAGE_CAPTION_BLOCK_LEN",
    "MAX_OPTION_TEXT_LEN",
    "MAX_QUESTION_TEXT_LEN",
    "SECTION_RANGES",
    "SHEET_NAME",
    "VALID_OPTIONS",
    "VALID_SECTIONS",
    "ExcelParser",
    "ParseError",
    "ParsedQuestion",
    "ParsedTest",
]

# ---------- DTOs ----------


@dataclass(frozen=True)
class ParseError:
    """One validation failure pinned to an Excel row (``line == 0`` for file-level errors)."""

    line: int
    message: str


@dataclass(frozen=True)
class ParsedQuestion:
    """One validated row from the uploaded test.

    ``has_image`` is the optional ``has_image`` column: when true the question
    is expected to carry an illustration, collected in-bot after upload. The
    image id is **not** part of the Excel contract — only the intent is.
    """

    section: str
    position: int
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_option: str
    has_image: bool = False


@dataclass(frozen=True)
class ParsedTest:
    """A full 50-question test ready for ``QuestionRepository.bulk_create``."""

    questions: tuple[ParsedQuestion, ...]


# ---------- Validation constants ----------
#
# The semantic limits (sections, lengths, caption budget) live in
# app/services/question_validation.py — shared with the web panel editor —
# and are re-exported above for backward compatibility. Only the
# Excel-file-specific constants remain here.

SHEET_NAME: Final[str] = "Questions"

# Recognised ``has_image`` tokens (case-insensitive, after stripping). Anything
# else is a parse error so a typo can't silently disable an illustration.
_HAS_IMAGE_TRUE: Final[frozenset[str]] = frozenset({"y", "yes", "true", "1", "да", "+", "✓", "x"})
_HAS_IMAGE_FALSE: Final[frozenset[str]] = frozenset({"", "n", "no", "false", "0", "нет", "-", "—"})

_REQUIRED_COLUMN_LABELS: Final[tuple[str, ...]] = (
    "section",
    "position",
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_option",
)

# Total columns we read positionally: the 8 required + the optional
# ``has_image`` in column I. Files authored before this column existed simply
# have it padded to None (→ falsy → text question), so they keep parsing.
_TOTAL_COLUMNS: Final[int] = 9


# ---------- Parser ----------


class ExcelParser:
    """Stateless parser. Constructed without arguments so a single instance can be reused."""

    def parse(self, file_bytes: bytes) -> ParsedTest | list[ParseError]:
        """Validate ``file_bytes`` and return either the parsed test or a list of errors."""
        try:
            workbook = load_workbook(
                BytesIO(file_bytes),
                read_only=True,
                data_only=True,
            )
        except Exception:
            # openpyxl can raise InvalidFileException, BadZipFile, OSError,
            # KeyError, ValueError and a few others — the parser's contract
            # is "never raises", so any failure here becomes a file-level
            # ParseError the admin can act on.
            return [ParseError(line=0, message="Не удалось прочитать файл. Проверьте формат.")]

        if SHEET_NAME not in workbook.sheetnames:
            return [
                ParseError(line=0, message=f"Лист «{SHEET_NAME}» не найден в файле."),
            ]

        worksheet = workbook[SHEET_NAME]
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))

        errors: list[ParseError] = []
        questions: list[ParsedQuestion] = []
        seen_positions: set[int] = set()
        section_counts: dict[str, int] = dict.fromkeys(VALID_SECTIONS, 0)

        if len(rows) != EXPECTED_ROW_COUNT:
            errors.append(
                ParseError(
                    line=0,
                    message=(f"Ожидалось {EXPECTED_ROW_COUNT} вопросов, найдено {len(rows)}."),
                )
            )

        for idx, raw_row in enumerate(rows, start=2):  # row 2 is the first data row
            # Pad/trim to the expected columns so missing trailing cells surface
            # as "пусто" errors rather than IndexError. ``has_image`` (col I) is
            # optional, so a short row pads it to None → text question.
            padded = tuple(raw_row) + (None,) * _TOTAL_COLUMNS
            cells = padded[:_TOTAL_COLUMNS]
            row_errors = _validate_row(cells, line=idx)

            if row_errors:
                errors.extend(row_errors)
                continue

            question = _build_question(cells)
            questions.append(question)

            # Duplicate-position detection happens here so we keep one error
            # per duplicate occurrence (line-referenced).
            if question.position in seen_positions:
                errors.append(
                    ParseError(
                        line=idx,
                        message=f"Позиция {question.position} уже использована выше.",
                    )
                )
            else:
                seen_positions.add(question.position)

            section_counts[question.section] += 1
            # Section ↔ position-range consistency is checked per-row inside
            # _validate_row (via validate_question_fields), so rows that fail
            # it never reach this point.

        # Per-section counts (skip if file-level row count already failed —
        # the counts will obviously also be wrong).
        if len(rows) == EXPECTED_ROW_COUNT:
            for section, (_, _, expected) in SECTION_RANGES.items():
                actual = section_counts[section]
                if actual != expected:
                    errors.append(
                        ParseError(
                            line=0,
                            message=(
                                f"Ожидалось {expected} вопросов в разделе "
                                f"«{section}», найдено {actual}."
                            ),
                        )
                    )

        if errors:
            return errors
        return ParsedTest(questions=tuple(questions))


# ---------- Row-level helpers ----------


def _validate_row(cells: tuple[object, ...], *, line: int) -> list[ParseError]:
    """Validate one row in isolation. Returns an empty list if all cells pass.

    Excel-specific concerns (cell presence, position typing, ``has_image``
    token parsing) are handled here; the semantic field rules are delegated
    to :mod:`app.services.question_validation` so the web panel shares them.
    """
    section, position, qtext, oa, ob, oc, od, correct, has_image_raw = cells
    required = cells[:8]
    errors: list[ParseError] = []

    # 1. Required cells present. ``has_image`` is optional and validated
    #    separately below, so it's excluded here.
    for label, value in zip(_REQUIRED_COLUMN_LABELS, required, strict=True):
        if value is None or (isinstance(value, str) and not value.strip()):
            errors.append(ParseError(line=line, message=f"Колонка «{label}» пустая."))
    if errors:
        return errors

    # 2. Position is an int (Excel-specific typing concern; range checks are
    #    semantic and delegated below).
    if not isinstance(position, int) or isinstance(position, bool):
        errors.append(
            ParseError(
                line=line,
                message="Колонка «position» должна быть целым числом.",
            )
        )
        return errors

    # 3. has_image token must be recognised (Excel-specific spelling concern).
    has_image, has_image_err = _parse_has_image(has_image_raw)
    if has_image_err is not None:
        errors.append(ParseError(line=line, message=has_image_err))

    # 4. Semantic field rules — shared with the web panel.
    field_errors = validate_question_fields(
        section=str(section).strip(),
        position=position,
        question_text=str(qtext).strip(),
        option_a=str(oa).strip(),
        option_b=str(ob).strip(),
        option_c=str(oc).strip(),
        option_d=str(od).strip(),
        correct_option=str(correct).strip().upper(),
        has_image=has_image,
    )
    errors.extend(ParseError(line=line, message=fe.message) for fe in field_errors)
    return errors


def _build_question(cells: tuple[object, ...]) -> ParsedQuestion:
    """Build a :class:`ParsedQuestion` from a row that already passed ``_validate_row``."""
    section, position, qtext, oa, ob, oc, od, correct, has_image_raw = cells
    # ``_validate_row`` already proved ``position`` is an int and the text
    # cells are stringable; the cast tells mypy what the runtime checks
    # already enforced.
    assert isinstance(position, int)
    has_image, _ = _parse_has_image(has_image_raw)
    return ParsedQuestion(
        section=str(section).strip(),
        position=position,
        question_text=str(qtext).strip(),
        option_a=str(oa).strip(),
        option_b=str(ob).strip(),
        option_c=str(oc).strip(),
        option_d=str(od).strip(),
        correct_option=str(correct).strip().upper(),
        has_image=has_image,
    )


def _parse_has_image(value: object) -> tuple[bool, str | None]:
    """Interpret the optional ``has_image`` cell → ``(flag, error_message)``.

    Accepts the usual yes/no spellings in Russian or English plus 1/0; an empty
    cell means "no image". An unrecognised token is an error so a typo never
    silently drops an illustration the teacher intended.
    """
    if value is None:
        return False, None
    # openpyxl hands TRUE/FALSE cells back as Python bools.
    if isinstance(value, bool):
        return value, None
    token = str(value).strip().lower()
    if token in _HAS_IMAGE_TRUE:
        return True, None
    if token in _HAS_IMAGE_FALSE:
        return False, None
    return False, "Колонка «has_image» должна быть пустой или одним из: да/нет, yes/no, 1/0."
